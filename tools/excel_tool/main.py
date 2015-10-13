def at(seq, i, default=None):
	return seq[i] if i < len(seq) else default
import pprint
pp = pprint.PrettyPrinter(indent=2)

import argparse
def parse_argv():
	root = argparse.ArgumentParser(prog='excel_tool', description='an excel tool')  
	verbs = root.add_subparsers(dest='verb')
	verbs.required = True

	new = verbs.add_parser('new', help='new an empty .xlsx file')
	new.add_argument('path', help='file path')

	transform = verbs.add_parser('transform', help='transform .xlsx files')
	transform.add_argument('files', nargs='+', metavar='file', help='file paths')

	meta = verbs.add_parser('meta', help='print the meta data')
	meta.add_argument('path', help='file path')

	# intransform
	# intransform and open
	# constraint

	return root.parse_args()

args = parse_argv()

from excel_tool.type import *
from excel_tool.index import *
from excel_tool.constraint import *
from excel_tool.metadata import MetaData


def get_struct(rootname, ws):
	def __metadecl_to_list(v):
		if not v:
			return []
		ret = eval(str(v))
		if type(ret) is tuple:
			return list(ret)
		else:
			return [ret]
	try:
		rowIt = ws.rows
		meta_row = rowIt.next()
		name_row = rowIt.next()
		return struct(rootname, *[list(name.value, *__metadecl_to_list(meta_row[i].value)) for i, name in enumerate(name_row)])
	except StopIteration:
		pass

def get_meta_data(ws):
	try:
		rowIt = ws.rows
		meta_row = rowIt.next()
		name_row = rowIt.next()
		return [ MetaData().init_from_metadecl(name.value, meta_row[i].value) for i, name in enumerate(name_row) ]
	except StopIteration:
		pass

def dump_src_file(name, meta):
	tpl = """
struct {struct_name}{{
  {field_decls}

  template<class Archive>
  void serialize(Archive& ar, const unsigned int version){{
    {serialize_exprs}
  }}

  {index_tags}
  typedef multi_index_container<
    {struct_name},
    indexed_by<
      {index_decls}
    >
  > map_type;

  template<class Tag> static auto get(){{ return data.get<Tag>(); }}
  static auto get(){{ return data.get<0>(); }}

  map_type data;
  static const char name[] = "{struct_name}";
}};
"""
	struct_name = name
	field_decls = "\n  ".join("{0} {1};".format(field.type.name, field.name) for field in meta)
	serialize_exprs = "\n    ".join("ar & {0};".format(field.name) for field in meta)

	def index_str(index, field):
		return "{0}<tag<i_{1}>, member<{2}, {3}, &{2}::{1}>>".format(index["tag"], field.name, struct_name, field.type.name)
	indice = [(index, field) for field in meta for index in field.indice]
	index_tags = "\n  ".join("struct i_{0}{{}};".format(field.name) for index, field in indice)
	index_decls = ",\n      ".join(index_str(index, field) for index, field in indice)

	print tpl.format(**locals())

import openpyxl
verb = args.verb
if verb == 'new':
	def f(path):
		wb = openpyxl.Workbook()
		wb.save(path)
	f(args.path)
elif verb == 'transform':
	def f(path):
		from os.path import splitext, basename
		(rootname, ext) = splitext(basename(path))

		wb = openpyxl.load_workbook(path, read_only=True)
		ws = wb.active
		meta = get_meta_data(ws)
		dump_src_file(rootname, meta)
	f(args.files[0])
elif verb == 'meta':
	def f(path):
		wb = openpyxl.load_workbook(path, read_only=True)
		ws = wb.active
		meta = get_meta_data(ws)
		pp.pprint([v.to_dict() for v in meta])
	f(args.path)

